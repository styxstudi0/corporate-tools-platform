"""Split an Excel workbook into separate files by worksheet."""

from __future__ import annotations

import argparse
import re
from pathlib import Path


def safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("_") or "sheet"


def split_excel(input_file: Path, output_dir: Path) -> list[Path]:
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError as exc:
        raise SystemExit("Missing dependency: pip install openpyxl") from exc

    workbook = load_workbook(input_file)
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    for sheet_name in workbook.sheetnames:
        source = workbook[sheet_name]
        new_workbook = Workbook()
        target = new_workbook.active
        target.title = sheet_name[:31]

        for row in source.iter_rows():
            for cell in row:
                target[cell.coordinate].value = cell.value

        output_file = output_dir / f"{safe_name(sheet_name)}.xlsx"
        new_workbook.save(output_file)
        written.append(output_file)

    return written


def main() -> None:
    parser = argparse.ArgumentParser(description="Split an Excel workbook into one file per worksheet.")
    parser.add_argument("input", type=Path, help="Path to the .xlsx workbook")
    parser.add_argument("--output-dir", type=Path, default=Path("excel_split_output"))
    args = parser.parse_args()

    files = split_excel(args.input, args.output_dir)
    for file in files:
        print(file)


if __name__ == "__main__":
    main()

