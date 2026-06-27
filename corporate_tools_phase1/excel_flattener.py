"""Flatten merged Excel cells into a regular tabular workbook."""

from __future__ import annotations

from pathlib import Path


def flatten_workbook(input_file: Path, output_file: Path) -> dict:
    from openpyxl import load_workbook

    workbook = load_workbook(input_file)
    merged_ranges = 0
    for sheet in workbook.worksheets:
        ranges = list(sheet.merged_cells.ranges)
        merged_ranges += len(ranges)
        for merged in ranges:
            value = sheet.cell(merged.min_row, merged.min_col).value
            sheet.unmerge_cells(str(merged))
            for row in sheet.iter_rows(min_row=merged.min_row, max_row=merged.max_row, min_col=merged.min_col, max_col=merged.max_col):
                for cell in row:
                    cell.value = value
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    return {"tool": "Excel Merge Flattener", "sheets": len(workbook.worksheets), "merged_ranges_flattened": merged_ranges, "output": str(output_file)}
