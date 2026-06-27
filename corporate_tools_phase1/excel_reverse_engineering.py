"""Explain Excel workbook formulas, dependencies, errors, and structure."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from automation_common import write_output


def analyze_workbook(workbook_path: Path) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Missing dependency: pip install openpyxl") from exc

    workbook = load_workbook(workbook_path, data_only=False)
    sheets = []
    formulas = []
    errors = []
    references = {}

    for sheet in workbook.worksheets:
        sheet_info = {"name": sheet.title, "max_row": sheet.max_row, "max_column": sheet.max_column}
        sheets.append(sheet_info)
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    refs = re.findall(r"(?:'[^']+'|[A-Za-z0-9_ ]+)?!?\$?[A-Z]{1,3}\$?\d+", value)
                    item = {
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "formula": value,
                        "references": refs,
                        "explanation": explain_formula(value),
                    }
                    formulas.append(item)
                    references[f"{sheet.title}!{cell.coordinate}"] = refs
                elif isinstance(value, str) and value.startswith("#"):
                    errors.append({"sheet": sheet.title, "cell": cell.coordinate, "error": value})

    return {
        "tool": "Excel Reverse Engineering",
        "workbook": workbook_path.name,
        "sheets": sheets,
        "formula_count": len(formulas),
        "formulas": formulas[:200],
        "errors": errors,
        "dependency_map": references,
        "recommendations": [
            "Move repeated formulas into named ranges or helper tables",
            "Document hidden lookup sheets and external links",
            "Add validation for divide-by-zero and missing lookup values",
        ],
    }


def explain_formula(formula: str) -> str:
    upper = formula.upper()
    parts = []
    if "VLOOKUP" in upper or "XLOOKUP" in upper:
        parts.append("lookup formula")
    if "IF(" in upper or "IFS(" in upper:
        parts.append("conditional logic")
    if "SUM" in upper:
        parts.append("aggregation")
    if "ERROR" in upper:
        parts.append("error handling")
    return ", ".join(parts) if parts else "calculation formula"


def main() -> None:
    parser = argparse.ArgumentParser(description="Reverse engineer Excel workbook logic.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=Path("excel_reverse_engineering.json"))
    parser.add_argument("--format", choices=["json", "md"], default="json")
    args = parser.parse_args()
    print(write_output(analyze_workbook(args.workbook), args.output, args.format))


if __name__ == "__main__":
    main()

