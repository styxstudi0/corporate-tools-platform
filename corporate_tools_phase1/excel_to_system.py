"""Convert spreadsheet workflows into databases, dashboards, and app plans."""

from __future__ import annotations

import argparse
from pathlib import Path
from automation_common import write_output


def plan_system(workbook_path: Path) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Missing dependency: pip install openpyxl") from exc
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    tables = []
    for sheet in workbook.worksheets:
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1), [])]
        tables.append({"table_name": sheet.title.lower().replace(" ", "_"), "source_sheet": sheet.title, "columns": headers})
    return {
        "tool": "Excel-to-System",
        "workbook": workbook_path.name,
        "proposed_database_tables": tables,
        "workflow_modules": ["intake", "review", "approval", "reporting", "admin"],
        "dashboard_metrics": ["open items", "monthly volume", "approval aging", "exceptions"],
        "build_plan": ["Import workbook", "Define schema", "Add roles", "Create workflows", "Build dashboards", "Deploy web app"],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert Excel workbook into a system blueprint.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=Path("excel_to_system.json"))
    parser.add_argument("--format", choices=["json", "md"], default="json")
    args = parser.parse_args()
    print(write_output(plan_system(args.workbook), args.output, args.format))


if __name__ == "__main__":
    main()

